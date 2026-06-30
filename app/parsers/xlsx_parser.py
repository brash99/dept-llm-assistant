from pathlib import Path

from openpyxl import load_workbook

from app.parser import Parser
from app.knowledge import document_from_text


class XLSXParser(Parser):

    name = "XLSXParser"
    supported_suffixes = {".xlsx"}

    def parse(self, path, root_path):

        path = Path(path)

        wb = load_workbook(path, data_only=True)

        sheets = []
        sheet_names = []
        sheet_lengths = []

        total_rows = 0
        total_columns = 0

        for ws in wb.worksheets:

            sheet_names.append(ws.title)

            lines = [f"=== Sheet: {ws.title} ==="]

            max_col = 0

            for row in ws.iter_rows(values_only=True):

                values = []

                for cell in row:
                    if cell is None:
                        values.append("")
                    else:
                        values.append(str(cell).strip())

                # Remove formatted-but-empty trailing cells
                while values and not values[-1].strip():
                    values.pop()

                # Skip completely empty rows
                if values and any(v.strip() for v in values):
                    lines.append(" | ".join(values))

                max_col = max(max_col, len(values))

            sheet_text = "\n".join(lines)

            sheets.append(sheet_text)
            sheet_lengths.append(len(sheet_text))

            total_rows += ws.max_row
            total_columns = max(total_columns, max_col)

        text = "\n\n".join(sheets).strip()

        props = wb.properties

        metadata = {
            "num_sheets": len(wb.sheetnames),
            "sheet_names": sheet_names,
            "sheet_lengths": sheet_lengths,
            "max_columns": total_columns,
            "total_rows": total_rows,
            "text_length": len(text),
            "quality": "good" if len(text) >= 100 else "poor",
            "is_empty": len(text) == 0,
            "core_properties": {
                "title": props.title,
                "subject": props.subject,
                "creator": props.creator,
                "keywords": props.keywords,
                "category": props.category,
                "created": props.created.isoformat() if props.created else None,
                "modified": props.modified.isoformat() if props.modified else None,
            },
        }

        title = props.title or path.stem

        return document_from_text(
            source_path=path,
            root_path=root_path,
            text=text,
            parser=self.name,
            title=title,
            metadata=metadata,
        )
