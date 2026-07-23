#!/usr/bin/env python3
"""Reconcile ISO three-year SCH with a bounded administration workbook extract."""

from __future__ import annotations

import argparse
import csv
from hashlib import sha256
import json
from pathlib import Path
import statistics
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.department_profiles import _repair_sch_rows, _schedule_row  # noqa: E402
from app.faculty_identity import FacultyIdentityService  # noqa: E402
from app.institutional_units import AcademicUnitRegistry  # noqa: E402
from app.metric_readiness_audit import load_normalized_objects  # noqa: E402
from app.reasoning.academic_unit_mapping import AcademicUnitMappingService  # noqa: E402
from scripts.build_department_sch_timeline import build_timeline  # noqa: E402


YEARS = ("2022-23", "2023-24", "2024-25")
ADMIN_TOTAL_SCH_COLUMN = 22  # V
ADMIN_LLC_SCH_COLUMN = 28  # AB


def read_administration_workbook(path, registry=None, sheet_name=None):
    registry = registry or AcademicUnitRegistry.load()
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    header_row, department_column, faculty_column = _find_headers(sheet)
    values = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        department = sheet.cell(row_number, department_column).value
        faculty = sheet.cell(row_number, faculty_column).value
        total_sch = sheet.cell(row_number, ADMIN_TOTAL_SCH_COLUMN).value
        llc_sch = sheet.cell(row_number, ADMIN_LLC_SCH_COLUMN).value
        if all(value in (None, "") for value in (department, faculty, total_sch, llc_sch)):
            continue
        if not department:
            raise ValueError(f"Administration row {row_number} has metrics without a department")
        unit = registry.resolve(str(department))
        if unit is None or not unit.is_department_workforce_unit:
            raise ValueError(
                f"Administration department does not resolve to a governed current "
                f"department at row {row_number}: {department}"
            )
        record = {
            "source_row": row_number,
            "published_department": str(department).strip(),
            "academic_unit_id": unit.unit_id,
            "department_name": unit.published_name,
            "admin_faculty": _number(faculty, "Faculty", row_number),
            "admin_three_year_average_sch": _number(
                total_sch, "Column V", row_number
            ),
            "admin_three_year_average_llc_sch": _number(
                llc_sch, "Column AB", row_number
            ),
        }
        values.append(record)
    ids = [item["academic_unit_id"] for item in values]
    if len(ids) != len(set(ids)):
        raise ValueError("Administration workbook contains duplicate governed departments")
    return tuple(sorted(values, key=lambda item: item["academic_unit_id"]))


def _find_headers(sheet):
    for row_number in range(1, min(sheet.max_row, 50) + 1):
        normalized = {
            column: " ".join(str(sheet.cell(row_number, column).value or "").split()).casefold()
            for column in range(1, sheet.max_column + 1)
        }
        departments = [
            column for column, value in normalized.items()
            if value in {"department", "academic department", "unit"}
        ]
        faculty = [
            column for column, value in normalized.items()
            if value in {"faculty", "faculty count", "number of faculty", "ft faculty"}
        ]
        if len(departments) == 1 and len(faculty) == 1:
            return row_number, departments[0], faculty[0]
    raise ValueError("Could not identify unique Department and Faculty headers")


def _number(value, field, row):
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"Administration {field} is not numeric at row {row}")
    return float(value)


def reconcile(total_timeline, llc_timeline, administration):
    total = {item["academic_unit_id"]: item for item in total_timeline["departments"]}
    llc = {item["academic_unit_id"]: item for item in llc_timeline["departments"]}
    admin = {item["academic_unit_id"]: item for item in administration}
    if set(total) != set(llc):
        raise ValueError("ISO total and LLC department populations differ")
    missing_admin = sorted(set(total) - set(admin))
    extra_admin = sorted(set(admin) - set(total))
    if missing_admin or extra_admin:
        raise ValueError(
            f"Administration/ISO department mismatch; missing={missing_admin}, "
            f"extra={extra_admin}"
        )
    rows = []
    for unit_id in sorted(total):
        total_years = _years(total[unit_id])
        llc_years = _years(llc[unit_id])
        iso_average = sum(total_years[year] for year in YEARS) / 3
        iso_llc_average = sum(llc_years[year] for year in YEARS) / 3
        source = admin[unit_id]
        sch_difference = iso_average - source["admin_three_year_average_sch"]
        llc_difference = (
            iso_llc_average - source["admin_three_year_average_llc_sch"]
        )
        faculty_difference = total[unit_id]["faculty"] - source["admin_faculty"]
        reasons = []
        if faculty_difference:
            reasons.append("different_faculty_denominator")
        if sch_difference:
            reasons.append("unknown_total_sch_difference")
        if llc_difference:
            reasons.append("unknown_llc_difference")
        rows.append({
            "academic_unit_id": unit_id,
            "department": total[unit_id]["department_name"],
            "admin_faculty": source["admin_faculty"],
            "iso_faculty": total[unit_id]["faculty"],
            "faculty_difference": faculty_difference,
            **{f"iso_ay_{year}_sch": total_years[year] for year in YEARS},
            "admin_three_year_average_sch": source[
                "admin_three_year_average_sch"
            ],
            "iso_three_year_average_sch": iso_average,
            "sch_difference": sch_difference,
            "sch_percent_difference": _percent(
                sch_difference, source["admin_three_year_average_sch"]
            ),
            **{f"iso_ay_{year}_llc_sch": llc_years[year] for year in YEARS},
            "admin_three_year_average_llc_sch": source[
                "admin_three_year_average_llc_sch"
            ],
            "iso_three_year_average_llc_sch": iso_llc_average,
            "llc_sch_difference": llc_difference,
            "llc_percent_difference": _percent(
                llc_difference, source["admin_three_year_average_llc_sch"]
            ),
            "explanation_categories": reasons or ["exact_numeric_agreement"],
            "explanation": _explanation(
                faculty_difference, sch_difference, llc_difference
            ),
            "admin_source_row": source["source_row"],
        })
    rows.sort(key=lambda item: (-abs(item["sch_difference"]), item["department"]))
    summary = _summary(rows)
    payload = {
        "academic_years": list(YEARS),
        "average_formula": "(AY 2022-23 + AY 2023-24 + AY 2024-25) / 3",
        "iso_llc_definition": (
            "SCH from schedule sections with a nonblank explicitly published "
            "Area of LLC value."
        ),
        "administration_columns": {
            "total_sch": "V",
            "llc_sch": "AB",
        },
        "rows": rows,
        "summary": summary,
    }
    payload["deterministic_fingerprint"] = sha256(
        json.dumps(payload, sort_keys=True, separators=(",", ":")).encode()
    ).hexdigest()
    return payload


def _years(department):
    values = {
        item["academic_year"]: float(item["sch"])
        for item in department["academic_years"]
    }
    missing = [year for year in YEARS if year not in values]
    if missing:
        raise ValueError(
            f"ISO department lacks required academic years "
            f"{department['academic_unit_id']}: {missing}"
        )
    return values


def _percent(difference, administration):
    return round(100 * difference / administration, 6) if administration else None


def _explanation(faculty_difference, sch_difference, llc_difference):
    notes = []
    if faculty_difference:
        notes.append(
            "The published administration and current ISO faculty counts differ; "
            "the source populations require institutional comparison."
        )
    if sch_difference:
        notes.append(
            "The extract contains only the administration average, so section-, "
            "ownership-, inclusion-, and academic-year causes cannot yet be "
            "distinguished."
        )
    if llc_difference:
        notes.append(
            "The administration LLC definition is not present in the bounded "
            "extract; the cause of the LLC difference is unknown."
        )
    return " ".join(notes) or "The extracted administration and ISO values agree exactly."


def _summary(rows):
    sch_abs = [abs(item["sch_difference"]) for item in rows]
    llc_abs = [abs(item["llc_sch_difference"]) for item in rows]
    percentages = [
        abs(item["sch_percent_difference"])
        for item in rows if item["sch_percent_difference"] is not None
    ]
    return {
        "department_count": len(rows),
        "mean_absolute_sch_difference": statistics.fmean(sch_abs) if sch_abs else 0,
        "median_absolute_sch_difference": statistics.median(sch_abs) if sch_abs else 0,
        "maximum_absolute_sch_difference": max(sch_abs, default=0),
        "mean_absolute_llc_difference": statistics.fmean(llc_abs) if llc_abs else 0,
        "faculty_exact_agreement_count": sum(
            item["faculty_difference"] == 0 for item in rows
        ),
        "sch_exact_agreement_count": sum(item["sch_difference"] == 0 for item in rows),
        **{
            f"sch_within_{threshold}_percent_count": sum(
                value <= threshold for value in percentages
            )
            for threshold in (1, 2, 5, 10)
        },
    }


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--admin-workbook", type=Path, required=True)
    parser.add_argument("--admin-sheet")
    parser.add_argument("--timeline", type=Path, required=True)
    parser.add_argument("--normalized-root", type=Path, default=Path("storage/normalized"))
    parser.add_argument("--workforce-output", type=Path, required=True)
    parser.add_argument("--profiles-output", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args(argv)
    total_timeline = json.loads(args.timeline.read_text(encoding="utf-8"))
    workforce_root = args.workforce_output.parent if args.workforce_output.is_file() else args.workforce_output
    profile_root = args.profiles_output.parent if args.profiles_output.is_file() else args.profiles_output
    decisions = _jsonl(workforce_root / "analytical_workforce_decisions.jsonl")
    profiles = _jsonl(profile_root / "department_profiles.jsonl")
    objects, integrity = load_normalized_objects(args.normalized_root)
    if integrity["invalid_json_count"]:
        raise ValueError("Normalized corpus contains invalid JSON")
    identities = FacultyIdentityService().audit(objects).identities
    schedule_identity = {
        source.knowledge_object_id: identity.identity_id
        for identity in identities
        for source in identity.source_observations
        if source.source_system == "schedule"
    }
    home = {
        item["faculty_identity_id"]: item["analytical_academic_unit_id"]
        for item in decisions if item["workforce_disposition"] == "include"
    }
    mapper = AcademicUnitMappingService()
    raw_rows = []
    for item in objects:
        if item.get("object_type") != "course_offering_observation":
            continue
        row = _schedule_row(item, schedule_identity, home, mapper)
        row["llc_area_raw"] = item.get("llc_area_raw")
        raw_rows.append(row)
    rows, _ = _repair_sch_rows(tuple(raw_rows))
    llc_rows = tuple(row for row in rows if str(row.get("llc_area_raw") or "").strip())
    llc_timeline = build_timeline(profiles, llc_rows)
    administration = read_administration_workbook(
        args.admin_workbook, sheet_name=args.admin_sheet
    )
    payload = reconcile(total_timeline, llc_timeline, administration)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    _write_csv(args.output_dir / "department_sch_reconciliation.csv", payload["rows"])
    _write_llc_csv(args.output_dir / "department_llc_reconciliation.csv", payload["rows"])
    (args.output_dir / "department_reconciliation_summary.json").write_text(
        json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    markdown = _markdown(payload)
    (args.output_dir / "department_sch_reconciliation.md").write_text(
        markdown, encoding="utf-8"
    )
    (args.output_dir / "department_reconciliation_summary.md").write_text(
        markdown, encoding="utf-8"
    )
    print(json.dumps({
        **payload["summary"],
        "deterministic_fingerprint": payload["deterministic_fingerprint"],
    }, indent=2, sort_keys=True))
    return 0


def _jsonl(path):
    return tuple(
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    )


def _write_csv(path, rows):
    fields = tuple(rows[0]) if rows else ()
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        if fields:
            writer.writeheader()
            for row in rows:
                writer.writerow({
                    key: json.dumps(value) if isinstance(value, list) else value
                    for key, value in row.items()
                })


def _write_llc_csv(path, rows):
    fields = (
        "academic_unit_id", "department", "admin_faculty", "iso_faculty",
        *tuple(f"iso_ay_{year}_llc_sch" for year in YEARS),
        "admin_three_year_average_llc_sch", "iso_three_year_average_llc_sch",
        "llc_sch_difference", "llc_percent_difference",
    )
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row[field] for field in fields})


def _markdown(payload):
    lines = [
        "# Administration SCH Reconciliation",
        "",
        f"- Academic years: {', '.join(payload['academic_years'])}",
        f"- Average: {payload['average_formula']}",
        f"- ISO LLC definition: {payload['iso_llc_definition']}",
        "",
        "## Summary",
        "",
    ]
    lines += [f"- {key}: {value}" for key, value in payload["summary"].items()]
    lines += [
        "",
        "## Department reconciliation",
        "",
        "| Department | Admin faculty | ISO faculty | Faculty Δ | Admin SCH | ISO SCH | SCH Δ | SCH Δ% | Admin LLC | ISO LLC | LLC Δ | LLC Δ% |",
        "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for item in payload["rows"]:
        lines.append(
            f"| {item['department']} | {item['admin_faculty']} | "
            f"{item['iso_faculty']} | {item['faculty_difference']} | "
            f"{item['admin_three_year_average_sch']} | "
            f"{item['iso_three_year_average_sch']} | "
            f"{item['sch_difference']} | {item['sch_percent_difference']} | "
            f"{item['admin_three_year_average_llc_sch']} | "
            f"{item['iso_three_year_average_llc_sch']} | "
            f"{item['llc_sch_difference']} | {item['llc_percent_difference']} |"
        )
    lines += ["", "## Difference explanations", ""]
    for item in payload["rows"]:
        if item["sch_difference"] or item["llc_sch_difference"] or item["faculty_difference"]:
            lines += [
                f"### {item['department']}",
                "",
                f"- Categories: {', '.join(item['explanation_categories'])}",
                f"- {item['explanation']}",
                "",
            ]
    return "\n".join(lines) + "\n"


if __name__ == "__main__":
    raise SystemExit(main())
