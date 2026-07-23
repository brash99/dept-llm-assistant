from __future__ import annotations

import json

from openpyxl import Workbook

from scripts.reconcile_administration_sch import (
    YEARS,
    _markdown,
    _write_csv,
    _write_llc_csv,
    read_administration_workbook,
    reconcile,
)


UNIT = "academic_unit:department_accounting_finance"
NAME = "Department of Accounting and Finance"


def _timeline(values, faculty=13):
    return {
        "departments": [{
            "academic_unit_id": UNIT,
            "department_name": NAME,
            "faculty": faculty,
            "academic_years": [
                {"academic_year": year, "sch": value}
                for year, value in zip(YEARS, values)
            ],
        }],
    }


def _admin(total=300, llc=100, faculty=13):
    return ({
        "source_row": 2,
        "published_department": "Accounting and Finance",
        "academic_unit_id": UNIT,
        "department_name": NAME,
        "admin_faculty": faculty,
        "admin_three_year_average_sch": total,
        "admin_three_year_average_llc_sch": llc,
    },)


def test_three_year_and_llc_averages_are_arithmetic_means():
    payload = reconcile(
        _timeline((150, 300, 450)),
        _timeline((50, 100, 150)),
        _admin(),
    )
    row = payload["rows"][0]
    assert row["iso_three_year_average_sch"] == 300
    assert row["iso_three_year_average_llc_sch"] == 100
    assert row["sch_difference"] == 0
    assert row["llc_sch_difference"] == 0


def test_faculty_and_metric_differences_are_signed_and_percentage_is_explicit():
    row = reconcile(
        _timeline((200, 200, 200), faculty=12),
        _timeline((50, 50, 50), faculty=12),
        _admin(total=100, llc=100, faculty=10),
    )["rows"][0]
    assert row["faculty_difference"] == 2
    assert row["sch_difference"] == 100
    assert row["sch_percent_difference"] == 100
    assert row["llc_sch_difference"] == -50
    assert "different_faculty_denominator" in row["explanation_categories"]


def test_workbook_extraction_reads_only_governed_fields_v_and_ab(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1, "Department")
    sheet.cell(1, 2, "Faculty")
    sheet.cell(1, 22, "Three-Year Average Total SCH")
    sheet.cell(1, 28, "Three-Year Average LLC SCH")
    sheet.cell(2, 1, "Accounting and Finance")
    sheet.cell(2, 2, 13)
    sheet.cell(2, 3, "ignored")
    sheet.cell(2, 22, 300)
    sheet.cell(2, 28, 100)
    path = tmp_path / "administration.xlsx"
    workbook.save(path)
    rows = read_administration_workbook(path)
    assert rows == _admin()


def test_reconciliation_is_deterministic_and_outputs_are_byte_identical(tmp_path):
    first = reconcile(
        _timeline((101, 202, 303)),
        _timeline((10, 20, 30)),
        _admin(total=200, llc=20),
    )
    second = reconcile(
        _timeline((101, 202, 303)),
        _timeline((10, 20, 30)),
        _admin(total=200, llc=20),
    )
    assert first == second
    assert json.dumps(first, sort_keys=True) == json.dumps(second, sort_keys=True)
    for number, payload in ((1, first), (2, second)):
        root = tmp_path / str(number)
        root.mkdir()
        _write_csv(root / "department_sch_reconciliation.csv", payload["rows"])
        _write_llc_csv(
            root / "department_llc_reconciliation.csv", payload["rows"]
        )
        (root / "summary.json").write_text(
            json.dumps(payload, indent=2, sort_keys=True) + "\n"
        )
        (root / "summary.md").write_text(_markdown(payload))
    assert {
        path.name: path.read_bytes() for path in (tmp_path / "1").iterdir()
    } == {
        path.name: path.read_bytes() for path in (tmp_path / "2").iterdir()
    }
